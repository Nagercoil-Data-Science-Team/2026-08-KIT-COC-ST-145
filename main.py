import os
import glob
import math
import pandas as pd
import numpy as np

from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MinMaxScaler

import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader, TensorDataset

import matplotlib.pyplot as plt

plt.rcParams["font.family"] = "Times New Roman"
plt.rcParams["font.weight"] = "bold"
plt.rcParams["font.size"] = 18
plt.rcParams["axes.labelweight"] = "bold"
plt.rcParams["axes.titleweight"] = "bold"
plt.rcParams["axes.grid"] = False


# ============================================================
# SETTINGS
# ============================================================

dataset_path = r"archive (2)"
MAX_ROWS_PER_FILE = 1000
RANDOM_STATE = 42


# ============================================================
# STEP 1: DATA LOADING
# ============================================================

print("=" * 70)
print("STEP 1: DATA LOADING")
print("=" * 70)

csv_files = sorted(glob.glob(os.path.join(dataset_path, "*.csv")))
print("CSV files found:", len(csv_files))

dataframes = []

for file in csv_files:
    filename = os.path.basename(file)
    print("\nLoading:", filename)

    try:
        df = pd.read_csv(file, nrows=MAX_ROWS_PER_FILE, low_memory=False)
        df.columns = df.columns.str.strip().str.replace("\ufeff", "", regex=False)
        df["Source_File"] = filename
        dataframes.append(df)
        print("Records loaded:", len(df))
        print("Columns:", len(df.columns))
    except Exception as e:
        print("ERROR while loading:", filename)
        print(e)


# ============================================================
# COMBINE DATA
# ============================================================

if len(dataframes) == 0:
    raise ValueError("No CSV files were successfully loaded.")

data = pd.concat(dataframes, ignore_index=True, sort=False)

print("\n" + "=" * 70)
print("DATA LOADING COMPLETED")
print("=" * 70)
print("Total files :", len(dataframes))
print("Total rows  :", len(data))
print("Total cols  :", len(data.columns))


# ============================================================
# STEP 2: DATA PREPROCESSING
# ============================================================

print("\n" + "=" * 70)
print("STEP 2: DATA PREPROCESSING")
print("=" * 70)


# ============================================================
# 2.1 REMOVE DUPLICATES
# ============================================================

before = len(data)
data = data.drop_duplicates()
after = len(data)

print("\n2.1 Duplicate removal")
print("Before:", before)
print("Removed:", before - after)
print("After:", after)


# ============================================================
# 2.2 CLEAN COLUMN NAMES
# ============================================================

data.columns = data.columns.str.strip().str.replace("\ufeff", "", regex=False)


# ============================================================
# 2.3 FIND LABEL COLUMN
# ============================================================

label_column = None

for col in data.columns:
    if col.strip().lower() == "label":
        label_column = col
        break

if label_column is None:
    raise ValueError("Label column was not found.")

print("\nLabel column:", label_column)


# ============================================================
# 2.4 REMOVE MISSING LABELS
# ============================================================

before = len(data)
data = data.dropna(subset=[label_column])
after = len(data)

print("Rows removed due to missing labels:", before - after)


# ============================================================
# 2.5 LABEL ENCODING
# ============================================================

print("\nOriginal label distribution:")
print(data[label_column].value_counts())

# Benign = 0, Any other label = Attack = 1
data["Target"] = data[label_column].apply(lambda x: 0 if str(x).strip().lower() == "benign" else 1)

# Remove original label
data.drop(columns=[label_column], inplace=True)

print("\nEncoded labels:")
print("Benign = 0")
print("Attack = 1")
print("\nEncoded distribution:")
print(data["Target"].value_counts())


# ============================================================
# 2.6 REMOVE IRRELEVANT COLUMNS
# ============================================================

irrelevant_columns = ["Flow ID", "FlowID", "Timestamp", "Source File", "Source_File"]
columns_to_remove = [col for col in irrelevant_columns if col in data.columns]

if columns_to_remove:
    data.drop(columns=columns_to_remove, inplace=True)
    print("\nRemoved irrelevant columns:", columns_to_remove)


# ============================================================
# 2.7 SEPARATE X AND Y
# ============================================================

X = data.drop(columns=["Target"])
y = data["Target"]


# ============================================================
# 2.8 CONVERT ALL FEATURES TO NUMERIC
# ============================================================

print("\nConverting features to numeric...")

for col in X.columns:
    X[col] = pd.to_numeric(X[col], errors="coerce")


# ============================================================
# 2.9 REMOVE INF / -INF
# ============================================================

print("\nChecking infinite values...")

X.replace([np.inf, -np.inf], np.nan, inplace=True)
print("Infinite values after replacement:", np.isinf(X.select_dtypes(include=np.number)).sum().sum())


# ============================================================
# 2.10 REMOVE EXTREMELY LARGE VALUES
# ============================================================

print("\nChecking extremely large values...")

X = X.astype(np.float64)
MAX_FLOAT = np.finfo(np.float64).max
X = X.mask(np.abs(X) > MAX_FLOAT)


# ============================================================
# 2.11 CHECK FINITE VALUES
# ============================================================

finite_check = np.isfinite(X.to_numpy())
non_finite_count = (~finite_check).sum()
print("Non-finite values found:", non_finite_count)

X = X.replace([np.inf, -np.inf], np.nan)


# ============================================================
# 2.12 REMOVE COMPLETELY EMPTY COLUMNS
# ============================================================

all_nan_columns = X.columns[X.isna().all()].tolist()

if all_nan_columns:
    print("\nRemoving completely empty columns:")
    print(all_nan_columns)
    X.drop(columns=all_nan_columns, inplace=True)


# ============================================================
# 2.13 HANDLE NaN VALUES
# ============================================================

print("\nHandling missing values...")

missing_before = X.isna().sum().sum()
print("Missing values before:", missing_before)

# Median imputation
for col in X.columns:
    median_value = X[col].median()
    if not np.isfinite(median_value):
        median_value = 0.0
    X[col] = X[col].fillna(median_value)

# Final safety replacement
X = X.fillna(0.0)


# ============================================================
# 2.14 FINAL FINITE VALUE CHECK
# ============================================================

X = X.replace([np.inf, -np.inf], 0.0)

X_values = X.to_numpy(dtype=np.float64)
X_values[~np.isfinite(X_values)] = 0.0
X = pd.DataFrame(X_values, columns=X.columns, index=X.index)

print("Missing values after:", X.isna().sum().sum())
print("Infinite values after:", np.isinf(X.to_numpy()).sum())
print("All values finite:", np.isfinite(X.to_numpy()).all())


# ============================================================
# STEP 3: FEATURE SELECTION (CORRELATION-BASED)
# ============================================================

print("\n" + "=" * 70)
print("STEP 3: FEATURE SELECTION")
print("=" * 70)

CORRELATION_THRESHOLD = 0.90

print("\nOriginal number of features:", X.shape[1])

corr_matrix = X.corr().abs()
upper_triangle = corr_matrix.where(np.triu(np.ones(corr_matrix.shape), k=1).astype(bool))

redundant_features = [col for col in upper_triangle.columns if any(upper_triangle[col] > CORRELATION_THRESHOLD)]

print("\nHighly correlated (redundant) features found:", len(redundant_features))
print(redundant_features)

X = X.drop(columns=redundant_features)

print("\nSelected number of features:", X.shape[1])


# ============================================================
# STEP 4: TRAIN / VALIDATION / TEST SPLIT
# ============================================================

print("\n" + "=" * 70)
print("STEP 4: DATASET SPLITTING")
print("=" * 70)

X_train_val, X_test, y_train_val, y_test = train_test_split(X, y, test_size=0.20, random_state=RANDOM_STATE, stratify=y)
X_train, X_val, y_train, y_val = train_test_split(X_train_val, y_train_val, test_size=0.20, random_state=RANDOM_STATE, stratify=y_train_val)

print("\nTraining records   :", len(X_train))
print("Validation records :", len(X_val))
print("Testing records    :", len(X_test))


# ============================================================
# STEP 4: MIN-MAX NORMALIZATION
# ============================================================

print("\n" + "=" * 70)
print("STEP 5: MIN-MAX NORMALIZATION")
print("=" * 70)

# Safety check BEFORE scaler
print("Train finite:", np.isfinite(X_train.to_numpy()).all())
print("Validation finite:", np.isfinite(X_val.to_numpy()).all())
print("Test finite:", np.isfinite(X_test.to_numpy()).all())

# Min-Max Scaler
scaler = MinMaxScaler(feature_range=(0, 1))

# Fit ONLY on training data
X_train_scaled = scaler.fit_transform(X_train)
X_val_scaled = scaler.transform(X_val)
X_test_scaled = scaler.transform(X_test)

# Convert back to DataFrame
X_train = pd.DataFrame(X_train_scaled, columns=X_train.columns, index=X_train.index)
X_val = pd.DataFrame(X_val_scaled, columns=X_val.columns, index=X_val.index)
X_test = pd.DataFrame(X_test_scaled, columns=X_test.columns, index=X_test.index)

# Final scaler safety check
X_train = X_train.replace([np.inf, -np.inf], 0.0).fillna(0.0)
X_val = X_val.replace([np.inf, -np.inf], 0.0).fillna(0.0)
X_test = X_test.replace([np.inf, -np.inf], 0.0).fillna(0.0)

print("\nMin-Max normalization completed.")


# ============================================================
# STEP 5: CLASS BALANCING
# ============================================================

print("\n" + "=" * 70)
print("STEP 6: CLASS BALANCING")
print("=" * 70)

print("\nBefore balancing:")
print(y_train.value_counts())

train_data = X_train.copy()
train_data["Target"] = y_train.values

# Separate classes
class_0 = train_data[train_data["Target"] == 0]
class_1 = train_data[train_data["Target"] == 1]

print("\nBenign:", len(class_0))
print("Attack:", len(class_1))

# Balance minority class
if len(class_0) != len(class_1):
    if len(class_0) > len(class_1):
        majority = class_0
        minority = class_1
    else:
        majority = class_1
        minority = class_0

    minority_oversampled = minority.sample(n=len(majority), replace=True, random_state=RANDOM_STATE)
    train_balanced = pd.concat([majority, minority_oversampled], ignore_index=True)

    # Shuffle
    train_balanced = train_balanced.sample(frac=1, random_state=RANDOM_STATE).reset_index(drop=True)
else:
    train_balanced = train_data.copy()

# Separate X and y
X_train = train_balanced.drop(columns=["Target"])
y_train = train_balanced["Target"]

print("\nAfter balancing:")
print(y_train.value_counts())


# ============================================================
# STEP 6: FINAL DATA INFORMATION
# ============================================================

print("\n" + "=" * 70)
print("STEP 7: FINAL PREPROCESSED DATA")
print("=" * 70)

print("\nX_train shape:", X_train.shape)
print("X_val shape  :", X_val.shape)
print("X_test shape :", X_test.shape)

print("\ny_train shape:", y_train.shape)
print("y_val shape  :", y_val.shape)
print("y_test shape :", y_test.shape)

print("\nNumber of features:", X_train.shape[1])


# ============================================================
# FINAL SAFETY CHECK
# ============================================================

print("\n" + "=" * 70)
print("STEP 8: FINAL NUMERICAL SAFETY CHECK")
print("=" * 70)

print("X_train contains NaN:", X_train.isna().any().any())
print("X_train contains Inf:", np.isinf(X_train.to_numpy()).any())
print("X_val contains NaN:", X_val.isna().any().any())
print("X_val contains Inf:", np.isinf(X_val.to_numpy()).any())
print("X_test contains NaN:", X_test.isna().any().any())
print("X_test contains Inf:", np.isinf(X_test.to_numpy()).any())


# ============================================================
# STEP 9: ADVERSARIAL SAMPLE GENERATION
# (Numerical-Aware TabTransformer + FGSM)
# ============================================================

print("\n" + "=" * 70)
print("STEP 9: ADVERSARIAL SAMPLE GENERATION")
print("=" * 70)

DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print("\nDevice:", DEVICE)

NUM_FEATURES = X_train.shape[1]
EMBED_DIM = 32
NUM_HEADS = 4
NUM_LAYERS = 2
FF_DIM = 64
DROPOUT = 0.1
BATCH_SIZE = 256
EPOCHS = 10
LEARNING_RATE = 1e-3
EPSILON = 0.05


# ============================================================
# 9.1 NUMERICAL-AWARE TABTRANSFORMER MODEL
# ============================================================

class NumericalTokenizer(nn.Module):

    def __init__(self, num_features, embed_dim):
        super().__init__()
        self.weight = nn.Parameter(torch.randn(num_features, embed_dim) * 0.02)
        self.bias = nn.Parameter(torch.zeros(num_features, embed_dim))

    def forward(self, x):
        return x.unsqueeze(-1) * self.weight + self.bias


class TabTransformer(nn.Module):

    def __init__(self, num_features, embed_dim, num_heads, num_layers, ff_dim, dropout):
        super().__init__()

        self.tokenizer = NumericalTokenizer(num_features, embed_dim)

        self.cls_token = nn.Parameter(torch.randn(1, 1, embed_dim) * 0.02)

        encoder_layer = nn.TransformerEncoderLayer(
            d_model=embed_dim,
            nhead=num_heads,
            dim_feedforward=ff_dim,
            dropout=dropout,
            batch_first=True
        )

        self.transformer = nn.TransformerEncoder(encoder_layer, num_layers=num_layers)

        self.norm = nn.LayerNorm(embed_dim)

        self.classifier = nn.Sequential(
            nn.Linear(embed_dim, ff_dim),
            nn.ReLU(),
            nn.Dropout(dropout),
            nn.Linear(ff_dim, 1)
        )

    def forward(self, x):
        tokens = self.tokenizer(x)
        batch_size = x.shape[0]
        cls_tokens = self.cls_token.expand(batch_size, -1, -1)
        tokens = torch.cat([cls_tokens, tokens], dim=1)
        encoded = self.transformer(tokens)
        cls_output = self.norm(encoded[:, 0, :])
        logits = self.classifier(cls_output)
        return logits.squeeze(-1)


# ============================================================
# 9.2 DATA PREPARATION FOR PYTORCH
# ============================================================

X_train_tensor = torch.tensor(X_train.to_numpy(), dtype=torch.float32)
y_train_tensor = torch.tensor(y_train.to_numpy(), dtype=torch.float32)

X_val_tensor = torch.tensor(X_val.to_numpy(), dtype=torch.float32)
y_val_tensor = torch.tensor(y_val.to_numpy(), dtype=torch.float32)

train_dataset = TensorDataset(X_train_tensor, y_train_tensor)
train_loader = DataLoader(train_dataset, batch_size=BATCH_SIZE, shuffle=True)

val_dataset = TensorDataset(X_val_tensor, y_val_tensor)
val_loader = DataLoader(val_dataset, batch_size=BATCH_SIZE, shuffle=False)


# ============================================================
# 9.3 TRAIN TABTRANSFORMER MODEL
# ============================================================

print("\nTraining TabTransformer model...")

model = TabTransformer(
    num_features=NUM_FEATURES,
    embed_dim=EMBED_DIM,
    num_heads=NUM_HEADS,
    num_layers=NUM_LAYERS,
    ff_dim=FF_DIM,
    dropout=DROPOUT
).to(DEVICE)

criterion = nn.BCEWithLogitsLoss()
optimizer = optim.Adam(model.parameters(), lr=LEARNING_RATE)

for epoch in range(EPOCHS):

    model.train()
    total_loss = 0.0

    for batch_X, batch_y in train_loader:

        batch_X = batch_X.to(DEVICE)
        batch_y = batch_y.to(DEVICE)

        optimizer.zero_grad()

        outputs = model(batch_X)
        loss = criterion(outputs, batch_y)

        loss.backward()
        optimizer.step()

        total_loss += loss.item() * batch_X.size(0)

    avg_train_loss = total_loss / len(train_dataset)

    model.eval()
    val_loss = 0.0
    correct = 0

    with torch.no_grad():

        for batch_X, batch_y in val_loader:

            batch_X = batch_X.to(DEVICE)
            batch_y = batch_y.to(DEVICE)

            outputs = model(batch_X)
            loss = criterion(outputs, batch_y)

            val_loss += loss.item() * batch_X.size(0)

            preds = (torch.sigmoid(outputs) > 0.5).float()
            correct += (preds == batch_y).sum().item()

    avg_val_loss = val_loss / len(val_dataset)
    val_accuracy = correct / len(val_dataset)

    print(
        "Epoch:", epoch + 1,
        "Train Loss:", round(avg_train_loss, 5),
        "Val Loss:", round(avg_val_loss, 5),
        "Val Accuracy:", round(val_accuracy, 5)
    )


# ============================================================
# 9.4 FGSM ADVERSARIAL SAMPLE GENERATION
# ============================================================

print("\nGenerating FGSM adversarial samples...")

model.eval()

# Feature-wise valid range (post Min-Max scaling)
FEATURE_MIN = 0.0
FEATURE_MAX = 1.0

X_test_tensor = torch.tensor(X_test.to_numpy(), dtype=torch.float32).to(DEVICE)
y_test_tensor = torch.tensor(y_test.to_numpy(), dtype=torch.float32).to(DEVICE)

X_test_tensor.requires_grad = True

outputs = model(X_test_tensor)
loss = criterion(outputs, y_test_tensor)

model.zero_grad()
loss.backward()

data_grad = X_test_tensor.grad.data

# FGSM perturbation: Adversarial Input = Original Input + eps * sign(gradient)
perturbation = EPSILON * data_grad.sign()

X_adv_tensor = X_test_tensor.detach() + perturbation

# Feature constraints: clip to physically meaningful (scaled) range
X_adv_tensor = torch.clamp(X_adv_tensor, FEATURE_MIN, FEATURE_MAX)


# ============================================================
# 9.5 EVALUATE MODEL ON ADVERSARIAL SAMPLES
# ============================================================

with torch.no_grad():

    clean_outputs = model(X_test_tensor.detach())
    clean_preds = (torch.sigmoid(clean_outputs) > 0.5).float()
    clean_accuracy = (clean_preds == y_test_tensor).float().mean().item()

    adv_outputs = model(X_adv_tensor)
    adv_preds = (torch.sigmoid(adv_outputs) > 0.5).float()
    adv_accuracy = (adv_preds == y_test_tensor).float().mean().item()

print("\nClean Test Accuracy:", round(clean_accuracy, 5))
print("Adversarial Test Accuracy:", round(adv_accuracy, 5))
print("Accuracy Drop Due to FGSM:", round(clean_accuracy - adv_accuracy, 5))


# ============================================================
# 9.6 SAVE ADVERSARIAL SAMPLES
# ============================================================

X_adv = pd.DataFrame(X_adv_tensor.detach().cpu().numpy(), columns=X_test.columns, index=X_test.index)
X_adv["Target"] = y_test.values

print("\nAdversarial samples generated:", X_adv.shape[0])
print("Adversarial sample feature count:", X_adv.shape[1] - 1)


# ============================================================
# STEP 10: COMBINE NORMAL AND ADVERSARIAL DATA
# ============================================================

print("\n" + "=" * 70)
print("STEP 10: COMBINE NORMAL AND ADVERSARIAL DATA")
print("=" * 70)

X_clean_all = pd.concat([X_train, X_val, X_test], ignore_index=True)
y_clean_all = pd.concat([y_train, y_val, y_test], ignore_index=True)

X_adv_features = X_adv.drop(columns=["Target"])
y_adv_labels = X_adv["Target"]

X_combined = pd.concat([X_clean_all, X_adv_features], ignore_index=True)
y_combined = pd.concat([y_clean_all, y_adv_labels], ignore_index=True)

print("\nClean samples      :", len(X_clean_all))
print("Adversarial samples:", len(X_adv_features))
print("Total combined     :", len(X_combined))

TEST_SIZE = 0.1
BATCH_SIZE = 32
EPOCHS = 30
LEARNING_RATE = 1e-3

X_train_c, X_test_c, y_train_c, y_test_c = train_test_split(
    X_combined, y_combined, test_size=TEST_SIZE, random_state=RANDOM_STATE, stratify=y_combined
)

X_train_c, X_val_c, y_train_c, y_val_c = train_test_split(
    X_train_c, y_train_c, test_size=TEST_SIZE, random_state=RANDOM_STATE, stratify=y_train_c
)

print("\nTrain records     :", len(X_train_c))
print("Validation records:", len(X_val_c))
print("Test records      :", len(X_test_c))

X_train_c_tensor = torch.tensor(X_train_c.to_numpy(), dtype=torch.float32)
y_train_c_tensor = torch.tensor(y_train_c.to_numpy(), dtype=torch.float32)

X_val_c_tensor = torch.tensor(X_val_c.to_numpy(), dtype=torch.float32).to(DEVICE)
y_val_c_tensor = torch.tensor(y_val_c.to_numpy(), dtype=torch.float32).to(DEVICE)

X_test_c_tensor = torch.tensor(X_test_c.to_numpy(), dtype=torch.float32).to(DEVICE)
y_test_c_tensor = torch.tensor(y_test_c.to_numpy(), dtype=torch.float32).to(DEVICE)

train_c_dataset = TensorDataset(X_train_c_tensor, y_train_c_tensor)
train_c_loader = DataLoader(train_c_dataset, batch_size=BATCH_SIZE, shuffle=True)


# ============================================================
# STEP 11: ADVERSARIAL INPUT PURIFICATION (DENOISING AUTOENCODER)
# ============================================================

print("\n" + "=" * 70)
print("STEP 11: ADVERSARIAL INPUT PURIFICATION (DAE)")
print("=" * 70)


class DenoisingAutoencoder(nn.Module):

    def __init__(self, input_dim, hidden_dim=64, latent_dim=32):
        super().__init__()

        self.encoder = nn.Sequential(
            nn.Linear(input_dim, hidden_dim),
            nn.ReLU(),
            nn.Linear(hidden_dim, latent_dim),
            nn.ReLU()
        )

        self.decoder = nn.Sequential(
            nn.Linear(latent_dim, hidden_dim),
            nn.ReLU(),
            nn.Linear(hidden_dim, input_dim),
            nn.Sigmoid()
        )

    def forward(self, x):
        z = self.encoder(x)
        out = self.decoder(z)
        return out


DAE_NOISE_STD = 0.05

dae = DenoisingAutoencoder(input_dim=NUM_FEATURES).to(DEVICE)
dae_criterion = nn.MSELoss()
dae_optimizer = optim.Adam(dae.parameters(), lr=LEARNING_RATE)

print("\nTraining Denoising Autoencoder...")

for epoch in range(EPOCHS):

    dae.train()
    total_loss = 0.0

    for batch_X, _ in train_c_loader:

        batch_X = batch_X.to(DEVICE)

        noise = torch.randn_like(batch_X) * DAE_NOISE_STD
        corrupted = torch.clamp(batch_X + noise, 0.0, 1.0)

        dae_optimizer.zero_grad()
        reconstructed = dae(corrupted)
        loss = dae_criterion(reconstructed, batch_X)
        loss.backward()
        dae_optimizer.step()

        total_loss += loss.item() * batch_X.size(0)

    avg_loss = total_loss / len(train_c_dataset)
    print("DAE Epoch:", epoch + 1, "Reconstruction Loss:", round(avg_loss, 6))

dae.eval()


# ============================================================
# STEP 11.1: ADVERSARIAL TRAINING OF TABTRANSFORMER ON COMBINED DATA
# ============================================================

print("\nTraining robust TabTransformer on combined (normal + adversarial) data...")

robust_model = TabTransformer(
    num_features=NUM_FEATURES,
    embed_dim=EMBED_DIM,
    num_heads=NUM_HEADS,
    num_layers=NUM_LAYERS,
    ff_dim=FF_DIM,
    dropout=DROPOUT
).to(DEVICE)

robust_criterion = nn.BCEWithLogitsLoss()
robust_optimizer = optim.Adam(robust_model.parameters(), lr=LEARNING_RATE)

val_c_dataset = TensorDataset(X_val_c_tensor, y_val_c_tensor)
val_c_loader = DataLoader(val_c_dataset, batch_size=BATCH_SIZE, shuffle=False)

train_loss_history = []
val_loss_history = []
train_acc_history = []
val_acc_history = []

for epoch in range(EPOCHS):

    robust_model.train()
    total_loss = 0.0
    train_correct = 0

    for batch_X, batch_y in train_c_loader:

        batch_X = batch_X.to(DEVICE)
        batch_y = batch_y.to(DEVICE)

        robust_optimizer.zero_grad()

        outputs = robust_model(batch_X)
        loss = robust_criterion(outputs, batch_y)

        loss.backward()
        robust_optimizer.step()

        total_loss += loss.item() * batch_X.size(0)

        train_preds = (torch.sigmoid(outputs) > 0.5).float()
        train_correct += (train_preds == batch_y).sum().item()

    avg_train_loss = total_loss / len(train_c_dataset)
    train_accuracy = train_correct / len(train_c_dataset)

    robust_model.eval()
    val_loss = 0.0
    correct = 0

    with torch.no_grad():

        for batch_X, batch_y in val_c_loader:

            batch_X = batch_X.to(DEVICE)
            batch_y = batch_y.to(DEVICE)

            outputs = robust_model(batch_X)
            loss = robust_criterion(outputs, batch_y)

            val_loss += loss.item() * batch_X.size(0)

            preds = (torch.sigmoid(outputs) > 0.5).float()
            correct += (preds == batch_y).sum().item()

    avg_val_loss = val_loss / len(val_c_dataset)
    val_accuracy = correct / len(val_c_dataset)

    train_loss_history.append(avg_train_loss)
    val_loss_history.append(avg_val_loss)
    train_acc_history.append(train_accuracy)
    val_acc_history.append(val_accuracy)

    print(
        "Robust Model Epoch:", epoch + 1,
        "Train Loss:", round(avg_train_loss, 5),
        "Train Accuracy:", round(train_accuracy, 5),
        "Val Loss:", round(avg_val_loss, 5),
        "Val Accuracy:", round(val_accuracy, 5)
    )


# ============================================================
# STEP 12: HHO OPTIMIZATION (HARRIS HAWKS OPTIMIZATION)
# ============================================================

print("\n" + "=" * 70)
print("STEP 12: HHO OPTIMIZATION")
print("=" * 70)

with torch.no_grad():
    X_val_denoised_tensor = dae(X_val_c_tensor)


def evaluate_fitness(params):

    alpha = float(np.clip(params[0], 0.0, 1.0))
    beta = float(np.clip(params[1], 0.0, 1.0))

    with torch.no_grad():

        blended_input = beta * X_val_denoised_tensor + (1 - beta) * X_val_c_tensor

        branch1_prob = torch.sigmoid(robust_model(X_val_c_tensor))
        branch2_prob = torch.sigmoid(robust_model(blended_input))

        final_prob = alpha * branch1_prob + (1 - alpha) * branch2_prob
        preds = (final_prob > 0.5).float()

        accuracy = (preds == y_val_c_tensor).float().mean().item()

    return -accuracy


def levy_flight(dim):

    beta_levy = 1.5

    sigma = (
        (math.gamma(1 + beta_levy) * math.sin(math.pi * beta_levy / 2)) /
        (math.gamma((1 + beta_levy) / 2) * beta_levy * (2 ** ((beta_levy - 1) / 2)))
    ) ** (1 / beta_levy)

    u = np.random.randn(dim) * sigma
    v = np.random.randn(dim)
    step = u / (np.abs(v) ** (1 / beta_levy))

    return step


DIM = 2
LB = np.array([0.0, 0.0])
UB = np.array([1.0, 1.0])
NUM_HAWKS = 10
MAX_ITER = 20

hawks = np.random.uniform(LB, UB, (NUM_HAWKS, DIM))

rabbit_position = np.zeros(DIM)
rabbit_fitness = float("inf")

for hawk in hawks:

    fitness = evaluate_fitness(hawk)

    if fitness < rabbit_fitness:
        rabbit_fitness = fitness
        rabbit_position = hawk.copy()

print("\nRunning Harris Hawks Optimization...")

HHO_CONVERGENCE_HISTORY = []

for t in range(MAX_ITER):

    for i in range(NUM_HAWKS):

        E1 = 2 * (1 - t / MAX_ITER)
        E0 = 2 * np.random.rand() - 1
        E = E1 * E0

        q = np.random.rand()

        if abs(E) >= 1:

            if q >= 0.5:
                rand_hawk = hawks[np.random.randint(0, NUM_HAWKS)]
                hawks[i] = rand_hawk - np.random.rand() * np.abs(rand_hawk - 2 * np.random.rand() * hawks[i])
            else:
                hawks[i] = (rabbit_position - hawks.mean(axis=0)) - np.random.rand() * (LB + np.random.rand() * (UB - LB))

        else:

            r = np.random.rand()

            if r >= 0.5 and abs(E) >= 0.5:
                J = 2 * (1 - np.random.rand())
                hawks[i] = (rabbit_position - hawks[i]) - E * np.abs(J * rabbit_position - hawks[i])

            elif r >= 0.5 and abs(E) < 0.5:
                hawks[i] = rabbit_position - E * np.abs(rabbit_position - hawks[i])

            elif r < 0.5 and abs(E) >= 0.5:
                J = 2 * (1 - np.random.rand())
                Y = rabbit_position - E * np.abs(J * rabbit_position - hawks[i])
                Z = Y + np.random.rand(DIM) * levy_flight(DIM)

                if evaluate_fitness(Y) < evaluate_fitness(hawks[i]):
                    hawks[i] = Y
                elif evaluate_fitness(Z) < evaluate_fitness(hawks[i]):
                    hawks[i] = Z

            else:
                J = 2 * (1 - np.random.rand())
                Y = rabbit_position - E * np.abs(J * rabbit_position - hawks.mean(axis=0))
                Z = Y + np.random.rand(DIM) * levy_flight(DIM)

                if evaluate_fitness(Y) < evaluate_fitness(hawks[i]):
                    hawks[i] = Y
                elif evaluate_fitness(Z) < evaluate_fitness(hawks[i]):
                    hawks[i] = Z

        hawks[i] = np.clip(hawks[i], LB, UB)

        fitness = evaluate_fitness(hawks[i])

        if fitness < rabbit_fitness:
            rabbit_fitness = fitness
            rabbit_position = hawks[i].copy()

    HHO_CONVERGENCE_HISTORY.append(-rabbit_fitness)

    print("HHO Iteration:", t + 1, "Best Validation Accuracy:", round(-rabbit_fitness, 5))

BEST_ALPHA = float(np.clip(rabbit_position[0], 0.0, 1.0))
BEST_BETA = float(np.clip(rabbit_position[1], 0.0, 1.0))

print("\nHHO Optimized Ensemble Weight (alpha):", round(BEST_ALPHA, 5))
print("HHO Optimized DAE Blend Factor (beta) :", round(BEST_BETA, 5))


# ============================================================
# STEP 13: HYBRID CLASSIFICATION (FINAL EVALUATION)
# ============================================================

print("\n" + "=" * 70)
print("STEP 13: HYBRID CLASSIFICATION")
print("=" * 70)

from sklearn.metrics import accuracy_score, confusion_matrix, classification_report

robust_model.eval()
dae.eval()

with torch.no_grad():

    # Branch 1: Original IDS branch (TabTransformer on raw input)
    branch1_prob = torch.sigmoid(robust_model(X_test_c_tensor))

    # Branch 2: DAE Defense branch (TabTransformer on denoised input)
    denoised_test = dae(X_test_c_tensor)
    blended_test = BEST_BETA * denoised_test + (1 - BEST_BETA) * X_test_c_tensor
    branch2_prob = torch.sigmoid(robust_model(blended_test))

    # Weighted fusion using HHO-optimized weights
    final_prob = BEST_ALPHA * branch1_prob + (1 - BEST_ALPHA) * branch2_prob
    final_preds = (final_prob > 0.5).float()

branch1_preds = (branch1_prob > 0.5).float()
branch2_preds = (branch2_prob > 0.5).float()

branch1_np = branch1_prob.cpu().numpy()
branch2_np = branch2_prob.cpu().numpy()
final_prob_np = final_prob.cpu().numpy()
final_preds_np = final_preds.cpu().numpy()
y_test_np = y_test_c_tensor.cpu().numpy()

print("\nSample-wise branch probabilities (first 10 test samples):")

for i in range(10):
    print(
        "Sample", i,
        "| Original Branch Prob:", round(float(branch1_np[i]), 4),
        "| DAE Branch Prob:", round(float(branch2_np[i]), 4),
        "| Final Prob:", round(float(final_prob_np[i]), 4),
        "| True Label:", int(y_test_np[i])
    )

branch1_accuracy = accuracy_score(y_test_np, branch1_preds.cpu().numpy())
branch2_accuracy = accuracy_score(y_test_np, branch2_preds.cpu().numpy())
final_accuracy = accuracy_score(y_test_np, final_preds_np)

print("\nOriginal Branch Accuracy      :", round(branch1_accuracy, 5))
print("DAE Defense Branch Accuracy   :", round(branch2_accuracy, 5))
print("Final Hybrid Classification Accuracy:", round(final_accuracy, 5))

conf_matrix = confusion_matrix(y_test_np, final_preds_np)

print("\nFinal Confusion Matrix:")
print(conf_matrix)

print("\nClassification Report:")
print(classification_report(y_test_np, final_preds_np, target_names=["Benign", "Attack"]))


# ============================================================
# STEP 14: EXTENDED EVALUATION METRICS
# ============================================================

print("\n" + "=" * 70)
print("STEP 14: EXTENDED EVALUATION METRICS")
print("=" * 70)

from sklearn.metrics import (
    precision_score,
    recall_score,
    f1_score,
    roc_auc_score,
    roc_curve,
    precision_recall_curve,
    average_precision_score
)
from sklearn.calibration import calibration_curve

branch1_preds_np = branch1_preds.cpu().numpy()
branch2_preds_np = branch2_preds.cpu().numpy()

MODEL_RESULTS = {}

for model_name, probs, preds in [
    ("Original Branch", branch1_np, branch1_preds_np),
    ("DAE Branch", branch2_np, branch2_preds_np),
    ("Final Hybrid", final_prob_np, final_preds_np),
]:

    acc = accuracy_score(y_test_np, preds)
    prec = precision_score(y_test_np, preds, zero_division=0)
    rec = recall_score(y_test_np, preds, zero_division=0)
    f1 = f1_score(y_test_np, preds, zero_division=0)
    auc = roc_auc_score(y_test_np, probs)

    cm = confusion_matrix(y_test_np, preds)
    tn, fp, fn, tp = cm.ravel()

    fpr_val = fp / (fp + tn) if (fp + tn) > 0 else 0.0
    fnr_val = fn / (fn + tp) if (fn + tp) > 0 else 0.0

    MODEL_RESULTS[model_name] = {
        "probs": probs,
        "preds": preds,
        "accuracy": acc,
        "precision": prec,
        "recall": rec,
        "f1": f1,
        "roc_auc": auc,
        "fpr": fpr_val,
        "fnr": fnr_val,
        "confusion_matrix": cm
    }

    print("\nModel:", model_name)
    print("Accuracy      :", round(acc, 5))
    print("Precision     :", round(prec, 5))
    print("Recall        :", round(rec, 5))
    print("F1-score      :", round(f1, 5))
    print("ROC-AUC       :", round(auc, 5))
    print("FPR           :", round(fpr_val, 5))
    print("FNR           :", round(fnr_val, 5))
    print("Confusion Matrix:")
    print(cm)


# ============================================================
# 14.1 ATTACK SUCCESS RATE (ASR)
# ============================================================

correct_clean_mask = (clean_preds == y_test_tensor)
flipped_mask = correct_clean_mask & (adv_preds != y_test_tensor)

num_correct_clean = correct_clean_mask.sum().item()
num_flipped = flipped_mask.sum().item()

ATTACK_SUCCESS_RATE = (num_flipped / num_correct_clean) if num_correct_clean > 0 else 0.0

print("\nAttack Success Rate (FGSM, undefended model):", round(ATTACK_SUCCESS_RATE, 5))


# ============================================================
# 14.2 ROBUST ACCURACY (DEFENDED HYBRID PIPELINE ON ADVERSARIAL DATA)
# ============================================================

with torch.no_grad():

    denoised_adv = dae(X_adv_tensor)
    blended_adv = BEST_BETA * denoised_adv + (1 - BEST_BETA) * X_adv_tensor

    branch1_adv_prob = torch.sigmoid(robust_model(X_adv_tensor))
    branch2_adv_prob = torch.sigmoid(robust_model(blended_adv))

    final_adv_prob = BEST_ALPHA * branch1_adv_prob + (1 - BEST_ALPHA) * branch2_adv_prob
    final_adv_preds = (final_adv_prob > 0.5).float()

    ROBUST_ACCURACY = (final_adv_preds == y_test_tensor).float().mean().item()

print("Robust Accuracy (defended hybrid pipeline on FGSM samples):", round(ROBUST_ACCURACY, 5))


# ============================================================
# STEP 15: BASELINE CNN AND LSTM MODELS (FOR COMPARISON)
# ============================================================

print("\n" + "=" * 70)
print("STEP 15: BASELINE CNN AND LSTM MODELS")
print("=" * 70)

BASELINE_EPOCHS = 10


class CNNBaseline(nn.Module):

    def __init__(self, num_features):
        super().__init__()

        self.conv1 = nn.Conv1d(1, 16, kernel_size=3, padding=1)
        self.conv2 = nn.Conv1d(16, 32, kernel_size=3, padding=1)
        self.pool = nn.AdaptiveAvgPool1d(1)
        self.fc = nn.Linear(32, 1)
        self.relu = nn.ReLU()

    def forward(self, x):
        x = x.unsqueeze(1)
        x = self.relu(self.conv1(x))
        x = self.relu(self.conv2(x))
        x = self.pool(x).squeeze(-1)
        out = self.fc(x)
        return out.squeeze(-1)


class LSTMBaseline(nn.Module):

    def __init__(self, num_features, hidden_dim=32):
        super().__init__()

        self.lstm = nn.LSTM(input_size=1, hidden_size=hidden_dim, num_layers=1, batch_first=True)
        self.fc = nn.Linear(hidden_dim, 1)

    def forward(self, x):
        x = x.unsqueeze(-1)
        _, (h_n, _) = self.lstm(x)
        out = self.fc(h_n[-1])
        return out.squeeze(-1)


def train_baseline_model(model_instance, loader, epochs, model_label):

    criterion_b = nn.BCEWithLogitsLoss()
    optimizer_b = optim.Adam(model_instance.parameters(), lr=1e-3)

    for epoch in range(epochs):

        model_instance.train()
        total_loss = 0.0

        for batch_X, batch_y in loader:

            batch_X = batch_X.to(DEVICE)
            batch_y = batch_y.to(DEVICE)

            optimizer_b.zero_grad()

            outputs = model_instance(batch_X)
            loss = criterion_b(outputs, batch_y)

            loss.backward()
            optimizer_b.step()

            total_loss += loss.item() * batch_X.size(0)

        avg_loss = total_loss / len(loader.dataset)
        print(model_label, "Epoch:", epoch + 1, "Loss:", round(avg_loss, 5))

    return model_instance


print("\nTraining CNN baseline model...")
cnn_model = CNNBaseline(NUM_FEATURES).to(DEVICE)
cnn_model = train_baseline_model(cnn_model, train_c_loader, BASELINE_EPOCHS, "CNN")
cnn_model.eval()

print("\nTraining LSTM baseline model...")
lstm_model = LSTMBaseline(NUM_FEATURES).to(DEVICE)
lstm_model = train_baseline_model(lstm_model, train_c_loader, BASELINE_EPOCHS, "LSTM")
lstm_model.eval()

with torch.no_grad():
    cnn_probs = torch.sigmoid(cnn_model(X_test_c_tensor)).cpu().numpy()
    lstm_probs = torch.sigmoid(lstm_model(X_test_c_tensor)).cpu().numpy()

cnn_auc = roc_auc_score(y_test_np, cnn_probs)
lstm_auc = roc_auc_score(y_test_np, lstm_probs)

print("\nCNN Test ROC-AUC :", round(cnn_auc, 5))
print("LSTM Test ROC-AUC:", round(lstm_auc, 5))


# ============================================================
# STEP 16: VISUALIZATION
# ============================================================

print("\n" + "=" * 70)
print("STEP 16: VISUALIZATION")
print("=" * 70)

import os

PLOT_DIR = "plots"
os.makedirs(PLOT_DIR, exist_ok=True)
PLOT_DPI = 800

DARK_COLORS = [
    "#1B4F72", "#943126", "#1E8449", "#B9770E",
    "#5B2C6F", "#117864", "#7B241C", "#154360"
]

FONT_SIZE = 18
FONT_PROPS = {"family": "Times New Roman", "weight": "bold", "size": FONT_SIZE}

hybrid_probs = MODEL_RESULTS["Final Hybrid"]["probs"]
hybrid_preds = MODEL_RESULTS["Final Hybrid"]["preds"]
hybrid_cm = MODEL_RESULTS["Final Hybrid"]["confusion_matrix"]

COMPARISON_MODELS = {
    "Proposed Model": hybrid_probs,
    "CNN": cnn_probs,
    "LSTM": lstm_probs
}


def annotate_bars(ax, bars):
    for bar in bars:
        height = bar.get_height()
        ax.annotate(
            str(round(height, 3)),
            xy=(bar.get_x() + bar.get_width() / 2, height),
            xytext=(0, 3),
            textcoords="offset points",
            ha="center",
            va="bottom",
            fontweight="bold",
            fontfamily="Times New Roman",
            fontsize=FONT_SIZE
        )


def style_axes(ax):
    ax.grid(False)
    ax.tick_params(axis="both", labelsize=FONT_SIZE)
    for label in ax.get_xticklabels() + ax.get_yticklabels():
        label.set_fontweight("bold")
        label.set_fontfamily("Times New Roman")


def save_plot(fig, filename):
    filepath = os.path.join(PLOT_DIR, filename)
    fig.savefig(filepath, dpi=PLOT_DPI, bbox_inches="tight")
    print("Saved:", filepath)


# ------------------------------------------------------------
# 16.1 CONFUSION MATRIX PLOT
# ------------------------------------------------------------

fig1, ax1 = plt.subplots(figsize=(10, 8))

im = ax1.imshow(hybrid_cm, cmap=plt.cm.Blues)

ax1.set_title("Confusion Matrix", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax1.set_xlabel("Predicted Label", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax1.set_ylabel("True Label", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax1.set_xticks([0, 1])
ax1.set_yticks([0, 1])
ax1.set_xticklabels(["Benign", "Attack"])
ax1.set_yticklabels(["Benign", "Attack"])

for i in range(hybrid_cm.shape[0]):
    for j in range(hybrid_cm.shape[1]):
        ax1.text(
            j, i, str(hybrid_cm[i, j]),
            ha="center", va="center",
            fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE,
            color="white" if hybrid_cm[i, j] > hybrid_cm.max() / 2 else "black"
        )

style_axes(ax1)
fig1.colorbar(im, ax=ax1)
save_plot(fig1, "01_confusion_matrix.png")


# ------------------------------------------------------------
# 16.2 MODEL ACCURACY PLOT (TRAIN VS VALIDATION)
# ------------------------------------------------------------

fig2, ax2 = plt.subplots(figsize=(14, 8))

epochs_range = list(range(1, EPOCHS + 1))
EPOCH_TICK_STEP = 2

ax2.plot(epochs_range, train_acc_history, color=DARK_COLORS[0], linewidth=2.5, label="Training Accuracy")
ax2.plot(epochs_range, val_acc_history, color=DARK_COLORS[1], linewidth=2.5, label="Validation Accuracy")

ax2.set_title("Model Accuracy - Training vs Validation", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax2.set_xlabel("Epoch", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax2.set_ylabel("Accuracy", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax2.set_xticks(epochs_range[::EPOCH_TICK_STEP])
ax2.set_xlim(epochs_range[0] - 0.5, epochs_range[-1] + 0.5)
style_axes(ax2)
ax2.legend(prop=FONT_PROPS)
save_plot(fig2, "02_model_accuracy.png")


# ------------------------------------------------------------
# 16.3 MODEL LOSS PLOT (TRAIN VS VALIDATION)
# ------------------------------------------------------------

fig3, ax3 = plt.subplots(figsize=(14, 8))

ax3.plot(epochs_range, train_loss_history, color=DARK_COLORS[2], linewidth=2.5, label="Training Loss")
ax3.plot(epochs_range, val_loss_history, color=DARK_COLORS[3], linewidth=2.5, label="Validation Loss")

ax3.set_title("Model Loss - Training vs Validation", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax3.set_xlabel("Epoch", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax3.set_ylabel("Loss", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax3.set_xticks(epochs_range[::EPOCH_TICK_STEP])
ax3.set_xlim(epochs_range[0] - 0.5, epochs_range[-1] + 0.5)
style_axes(ax3)
ax3.legend(prop=FONT_PROPS)
save_plot(fig3, "03_model_loss.png")


# ------------------------------------------------------------
# 16.4 ROC CURVE COMPARISON (PROPOSED MODEL VS CNN VS LSTM)
# ------------------------------------------------------------

fig4, ax4 = plt.subplots(figsize=(10, 8))

for idx, (label, probs) in enumerate(COMPARISON_MODELS.items()):

    fpr_curve, tpr_curve, _ = roc_curve(y_test_np, probs)
    auc_val = roc_auc_score(y_test_np, probs)

    ax4.plot(
        fpr_curve, tpr_curve,
        color=DARK_COLORS[idx % len(DARK_COLORS)],
        linewidth=2.5,
        label=label + " (AUC = " + str(round(auc_val, 3)) + ")"
    )

ax4.plot([0, 1], [0, 1], color="#808080", linewidth=1.5, linestyle="--", label="Random Classifier")

ax4.set_title("ROC Curve Comparison", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax4.set_xlabel("False Positive Rate", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax4.set_ylabel("True Positive Rate", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
style_axes(ax4)
ax4.legend(prop=FONT_PROPS)
save_plot(fig4, "04_roc_curve_comparison.png")


# ------------------------------------------------------------
# 16.5 PRECISION-RECALL CURVE COMPARISON (PROPOSED MODEL VS CNN VS LSTM)
# ------------------------------------------------------------

fig5, ax5 = plt.subplots(figsize=(10, 8))

for idx, (label, probs) in enumerate(COMPARISON_MODELS.items()):

    precision_curve, recall_curve, _ = precision_recall_curve(y_test_np, probs)
    ap_val = average_precision_score(y_test_np, probs)

    ax5.plot(
        recall_curve, precision_curve,
        color=DARK_COLORS[idx % len(DARK_COLORS)],
        linewidth=2.5,
        label=label + " (AP = " + str(round(ap_val, 3)) + ")"
    )

ax5.set_title("Precision-Recall Curve Comparison", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax5.set_xlabel("Recall", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax5.set_ylabel("Precision", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
style_axes(ax5)
ax5.legend(prop=FONT_PROPS)
save_plot(fig5, "05_precision_recall_curve_comparison.png")


# ------------------------------------------------------------
# 16.6 PRECISION AND RECALL VS THRESHOLD (PROPOSED MODEL, SEPARATE COLORED LINES)
# ------------------------------------------------------------

fig5b, ax5b = plt.subplots(figsize=(10, 8))

precision_vals, recall_vals, threshold_vals = precision_recall_curve(y_test_np, hybrid_probs)

ax5b.plot(threshold_vals, precision_vals[:-1], color=DARK_COLORS[0], linewidth=2.5, label="Precision")
ax5b.plot(threshold_vals, recall_vals[:-1], color=DARK_COLORS[4], linewidth=2.5, label="Recall")

ax5b.set_title("Precision and Recall vs Threshold", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax5b.set_xlabel("Threshold", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax5b.set_ylabel("Score", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
style_axes(ax5b)
ax5b.legend(prop=FONT_PROPS)
save_plot(fig5b, "06_precision_recall_vs_threshold.png")


# ------------------------------------------------------------
# 16.7 CALIBRATION CURVE
# ------------------------------------------------------------

fig6, ax6 = plt.subplots(figsize=(10, 8))

prob_true, prob_pred = calibration_curve(y_test_np, hybrid_probs, n_bins=10)

ax6.plot(prob_pred, prob_true, marker="o", color=DARK_COLORS[5], linewidth=2.5, markersize=8, label="Calibration Curve")
ax6.plot([0, 1], [0, 1], color="#808080", linewidth=1.5, linestyle="--", label="Perfectly Calibrated")

ax6.set_title("Calibration Curve", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax6.set_xlabel("Mean Predicted Probability", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax6.set_ylabel("Fraction of Positives", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
style_axes(ax6)
ax6.legend(prop=FONT_PROPS)
save_plot(fig6, "07_calibration_curve.png")


# ------------------------------------------------------------
# 16.8 FPR AND FNR BAR PLOT
# ------------------------------------------------------------

fig7, ax7 = plt.subplots(figsize=(10, 8))

fpr_fnr_labels = ["FPR", "FNR"]
fpr_fnr_values = [MODEL_RESULTS["Final Hybrid"]["fpr"], MODEL_RESULTS["Final Hybrid"]["fnr"]]

bars_rate = ax7.bar(fpr_fnr_labels, fpr_fnr_values, width=0.4, color=[DARK_COLORS[6], DARK_COLORS[7]])

annotate_bars(ax7, bars_rate)

ax7.set_title("False Positive Rate vs False Negative Rate", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax7.set_xlabel("Metric", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax7.set_ylabel("Rate", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
style_axes(ax7)
save_plot(fig7, "08_fpr_fnr_bar_plot.png")


# ------------------------------------------------------------
# 16.9 PERFORMANCE METRICS BAR PLOT
# ------------------------------------------------------------

fig8, ax8 = plt.subplots(figsize=(13, 9))

metric_names = ["accuracy", "precision", "recall", "f1"]
metric_labels = ["Accuracy", "Precision", "Recall", "F1-score"]
metric_bar_colors = [DARK_COLORS[0], DARK_COLORS[2], DARK_COLORS[4], DARK_COLORS[6]]

metric_values = [MODEL_RESULTS["Final Hybrid"][m] for m in metric_names]

bars_metrics = ax8.bar(metric_labels, metric_values, width=0.6, color=metric_bar_colors)

annotate_bars(ax8, bars_metrics)

ax8.set_title("Performance Metrics", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax8.set_xlabel("Metric", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax8.set_ylabel("Score", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax8.set_ylim(0, 1.15)
style_axes(ax8)
save_plot(fig8, "09_performance_metrics_bar_plot.png")


# ------------------------------------------------------------
# 16.10 ATTACK SUCCESS RATE PLOT (SEPARATE WINDOW)
# ------------------------------------------------------------

fig9, ax9 = plt.subplots(figsize=(10, 8))

bars_asr = ax9.bar(["Attack Success Rate"], [ATTACK_SUCCESS_RATE], color=DARK_COLORS[6], width=0.4)

annotate_bars(ax9, bars_asr)

ax9.set_title("Attack Success Rate", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax9.set_xlabel("Metric", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax9.set_ylabel("Rate", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax9.set_ylim(0, 1)
style_axes(ax9)
save_plot(fig9, "10_attack_success_rate.png")


# ------------------------------------------------------------
# 16.11 ROBUST ACCURACY PLOT (SEPARATE WINDOW)
# ------------------------------------------------------------

fig10, ax10 = plt.subplots(figsize=(10, 8))

bars_robust = ax10.bar(["Robust Accuracy"], [ROBUST_ACCURACY], color=DARK_COLORS[7], width=0.4)

annotate_bars(ax10, bars_robust)

ax10.set_title("Robust Accuracy", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax10.set_xlabel("Metric", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax10.set_ylabel("Accuracy", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax10.set_ylim(0, 1)
style_axes(ax10)
save_plot(fig10, "11_robust_accuracy.png")

print("\nAll plots saved to folder:", os.path.abspath(PLOT_DIR))


# ============================================================
# STEP 17: FEATURE SELECTION RESULTS - REPORT + FIGURE
# ============================================================

print("\n" + "=" * 70)
print("STEP 17: FEATURE SELECTION RESULTS")
print("=" * 70)

N_ORIGINAL_FEATURES = corr_matrix.shape[0]
N_SELECTED_FEATURES = X.shape[1]
N_REMOVED_FEATURES = len(redundant_features)
PCT_FEATURE_REDUCTION = (N_REMOVED_FEATURES / N_ORIGINAL_FEATURES) * 100

print("\nOriginal features   :", N_ORIGINAL_FEATURES)
print("Selected features   :", N_SELECTED_FEATURES)
print("Removed features    :", N_REMOVED_FEATURES)
print("Feature reduction(%):", round(PCT_FEATURE_REDUCTION, 2))

N_CORR_FEATURES = corr_matrix.shape[0]
HEATMAP_FONT_SIZE = 6
HEATMAP_FIGSIZE = (max(14, N_CORR_FEATURES * 0.45), max(12, N_CORR_FEATURES * 0.45))

fig11, ax11 = plt.subplots(figsize=HEATMAP_FIGSIZE)
im11 = ax11.imshow(corr_matrix, cmap="coolwarm", vmin=-1, vmax=1)

ax11.set_title("Feature Correlation Heatmap", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)

ax11.set_xticks(np.arange(N_CORR_FEATURES))
ax11.set_yticks(np.arange(N_CORR_FEATURES))
ax11.set_xticklabels(corr_matrix.columns, fontsize=HEATMAP_FONT_SIZE, fontfamily="Times New Roman", fontweight="bold", rotation=90)
ax11.set_yticklabels(corr_matrix.columns, fontsize=HEATMAP_FONT_SIZE, fontfamily="Times New Roman", fontweight="bold")

for i in range(N_CORR_FEATURES):
    for j in range(N_CORR_FEATURES):
        corr_val = corr_matrix.iloc[i, j]
        ax11.text(
            j, i, str(round(corr_val, 2)),
            ha="center", va="center",
            fontsize=HEATMAP_FONT_SIZE, fontfamily="Times New Roman", fontweight="bold",
            color="white" if abs(corr_val) > 0.6 else "black"
        )

fig11.colorbar(im11, ax=ax11)
fig11.tight_layout()
save_plot(fig11, "12_correlation_heatmap.png")


# ============================================================
# STEP 18: FGSM ATTACK IMPACT (BASELINE MODEL) - REPORT + FIGURE
# ============================================================

print("\n" + "=" * 70)
print("STEP 18: FGSM ATTACK IMPACT (BASELINE MODEL)")
print("=" * 70)

clean_probs_np = torch.sigmoid(clean_outputs).detach().cpu().numpy()
adv_probs_np = torch.sigmoid(adv_outputs).detach().cpu().numpy()
clean_preds_np = clean_preds.detach().cpu().numpy()
adv_preds_np = adv_preds.detach().cpu().numpy()
y_test_orig_np = y_test_tensor.detach().cpu().numpy()

baseline_clean_metrics = {
    "accuracy": accuracy_score(y_test_orig_np, clean_preds_np),
    "precision": precision_score(y_test_orig_np, clean_preds_np, zero_division=0),
    "recall": recall_score(y_test_orig_np, clean_preds_np, zero_division=0),
    "f1": f1_score(y_test_orig_np, clean_preds_np, zero_division=0),
    "roc_auc": roc_auc_score(y_test_orig_np, clean_probs_np),
}

baseline_fgsm_metrics = {
    "accuracy": accuracy_score(y_test_orig_np, adv_preds_np),
    "precision": precision_score(y_test_orig_np, adv_preds_np, zero_division=0),
    "recall": recall_score(y_test_orig_np, adv_preds_np, zero_division=0),
    "f1": f1_score(y_test_orig_np, adv_preds_np, zero_division=0),
    "roc_auc": roc_auc_score(y_test_orig_np, adv_probs_np),
}

BASELINE_ACCURACY_DEGRADATION = baseline_clean_metrics["accuracy"] - baseline_fgsm_metrics["accuracy"]
BASELINE_ROBUST_ACCURACY = baseline_fgsm_metrics["accuracy"]

print("\nBaseline - Clean Test Performance:")
for k, v in baseline_clean_metrics.items():
    print(k, ":", round(v, 5))

print("\nBaseline - FGSM Adversarial Performance:")
for k, v in baseline_fgsm_metrics.items():
    print(k, ":", round(v, 5))

print("\nAttack Success Rate (ASR)     :", round(ATTACK_SUCCESS_RATE, 5))
print("Robust Accuracy (undefended)  :", round(BASELINE_ROBUST_ACCURACY, 5))
print("Accuracy Degradation          :", round(BASELINE_ACCURACY_DEGRADATION, 5))

fig12, ax12 = plt.subplots(figsize=(12, 8))

metric_labels_18 = ["Accuracy", "Precision", "Recall", "F1-score", "ROC-AUC"]
metric_keys_18 = ["accuracy", "precision", "recall", "f1", "roc_auc"]

x_pos = np.arange(len(metric_labels_18))
bar_width = 0.35

clean_vals_18 = [baseline_clean_metrics[k] for k in metric_keys_18]
fgsm_vals_18 = [baseline_fgsm_metrics[k] for k in metric_keys_18]

bars_clean = ax12.bar(x_pos - bar_width / 2, clean_vals_18, bar_width, color=DARK_COLORS[0], label="Clean")
bars_fgsm = ax12.bar(x_pos + bar_width / 2, fgsm_vals_18, bar_width, color=DARK_COLORS[1], label="FGSM")

annotate_bars(ax12, bars_clean)
annotate_bars(ax12, bars_fgsm)

ax12.set_title("Clean vs FGSM Performance (Baseline Model)", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax12.set_xlabel("Metric", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax12.set_ylabel("Score", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax12.set_xticks(x_pos)
ax12.set_xticklabels(metric_labels_18)
style_axes(ax12)
ax12.legend(prop=FONT_PROPS)
save_plot(fig12, "13_clean_vs_fgsm_performance.png")


# ============================================================
# STEP 19: ADVERSARIAL TRAINING EFFECT - REPORT ONLY
# ============================================================

print("\n" + "=" * 70)
print("STEP 19: ADVERSARIAL TRAINING EFFECT")
print("=" * 70)

with torch.no_grad():
    robust_clean_outputs = robust_model(X_test_tensor.detach())
    robust_clean_probs_t = torch.sigmoid(robust_clean_outputs)
    robust_clean_preds_t = (robust_clean_probs_t > 0.5).float()

robust_clean_correct_mask = (robust_clean_preds_t == y_test_tensor)

branch1_adv_preds_t = (branch1_adv_prob > 0.5).float()
flipped_noDAE_mask = robust_clean_correct_mask & (branch1_adv_preds_t != y_test_tensor)

num_robust_clean_correct = robust_clean_correct_mask.sum().item()
num_flipped_noDAE = flipped_noDAE_mask.sum().item()

ASR_WITH_ADV_TRAINING = (num_flipped_noDAE / num_robust_clean_correct) if num_robust_clean_correct > 0 else 0.0

branch1_adv_probs_np = branch1_adv_prob.detach().cpu().numpy()
branch1_adv_preds_np = branch1_adv_preds_t.detach().cpu().numpy()

with_adv_training_metrics = {
    "robust_accuracy": accuracy_score(y_test_orig_np, branch1_adv_preds_np),
    "f1": f1_score(y_test_orig_np, branch1_adv_preds_np, zero_division=0),
    "recall": recall_score(y_test_orig_np, branch1_adv_preds_np, zero_division=0),
    "roc_auc": roc_auc_score(y_test_orig_np, branch1_adv_probs_np),
    "asr": ASR_WITH_ADV_TRAINING,
}

without_adv_training_metrics = {
    "robust_accuracy": BASELINE_ROBUST_ACCURACY,
    "f1": baseline_fgsm_metrics["f1"],
    "recall": baseline_fgsm_metrics["recall"],
    "roc_auc": baseline_fgsm_metrics["roc_auc"],
    "asr": ATTACK_SUCCESS_RATE,
}

adv_training_comparison = pd.DataFrame(
    {
        "Without Adversarial Training": without_adv_training_metrics,
        "With Adversarial Training": with_adv_training_metrics,
    }
)

print("\n", adv_training_comparison.round(5))


# ============================================================
# STEP 20: DAE DEFENSE EFFECT - REPORT + FIGURE
# ============================================================

print("\n" + "=" * 70)
print("STEP 20: DAE DEFENSE EFFECT")
print("=" * 70)

branch2_adv_preds_t = (branch2_adv_prob > 0.5).float()
flipped_withDAE_mask = robust_clean_correct_mask & (branch2_adv_preds_t != y_test_tensor)
num_flipped_withDAE = flipped_withDAE_mask.sum().item()
ASR_WITH_DAE = (num_flipped_withDAE / num_robust_clean_correct) if num_robust_clean_correct > 0 else 0.0

branch2_adv_probs_np = branch2_adv_prob.detach().cpu().numpy()
branch2_adv_preds_np = branch2_adv_preds_t.detach().cpu().numpy()

without_dae_metrics = {
    "accuracy": accuracy_score(y_test_orig_np, branch1_adv_preds_np),
    "precision": precision_score(y_test_orig_np, branch1_adv_preds_np, zero_division=0),
    "recall": recall_score(y_test_orig_np, branch1_adv_preds_np, zero_division=0),
    "f1": f1_score(y_test_orig_np, branch1_adv_preds_np, zero_division=0),
    "robust_accuracy": accuracy_score(y_test_orig_np, branch1_adv_preds_np),
    "asr": ASR_WITH_ADV_TRAINING,
}

with_dae_metrics = {
    "accuracy": accuracy_score(y_test_orig_np, branch2_adv_preds_np),
    "precision": precision_score(y_test_orig_np, branch2_adv_preds_np, zero_division=0),
    "recall": recall_score(y_test_orig_np, branch2_adv_preds_np, zero_division=0),
    "f1": f1_score(y_test_orig_np, branch2_adv_preds_np, zero_division=0),
    "robust_accuracy": accuracy_score(y_test_orig_np, branch2_adv_preds_np),
    "asr": ASR_WITH_DAE,
}

DAE_ACCURACY_CHANGE = with_dae_metrics["accuracy"] - without_dae_metrics["accuracy"]

print("\nFGSM Samples WITHOUT DAE:")
for k, v in without_dae_metrics.items():
    print(k, ":", round(v, 5))

print("\nFGSM Samples WITH DAE:")
for k, v in with_dae_metrics.items():
    print(k, ":", round(v, 5))

print("\nAccuracy Change Due to DAE:", round(DAE_ACCURACY_CHANGE, 5))

def annotate_bars_rotated(ax, bars):
    for bar in bars:
        height = bar.get_height()
        ax.annotate(
            str(round(height, 3)),
            xy=(bar.get_x() + bar.get_width() / 2, height),
            xytext=(0, 8),
            textcoords="offset points",
            ha="center",
            va="bottom",
            rotation=90,
            fontweight="bold",
            fontfamily="Times New Roman",
            fontsize=FONT_SIZE
        )


fig13, ax13 = plt.subplots(figsize=(16, 10))

metric_labels_20 = ["Accuracy", "Precision", "Recall", "F1-score", "Robust Acc."]
metric_keys_20 = ["accuracy", "precision", "recall", "f1", "robust_accuracy"]

x_pos20 = np.arange(len(metric_labels_20))

without_vals_20 = [without_dae_metrics[k] for k in metric_keys_20]
with_vals_20 = [with_dae_metrics[k] for k in metric_keys_20]

bars_without = ax13.bar(x_pos20 - bar_width / 2, without_vals_20, bar_width, color=DARK_COLORS[2], label="Without DAE")
bars_with = ax13.bar(x_pos20 + bar_width / 2, with_vals_20, bar_width, color=DARK_COLORS[3], label="With DAE")

annotate_bars_rotated(ax13, bars_without)
annotate_bars_rotated(ax13, bars_with)

ax13.set_title("Performance Before vs After DAE Purification", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax13.set_xlabel("Metric", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax13.set_ylabel("Score", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax13.set_xticks(x_pos20)
ax13.set_xticklabels(metric_labels_20)
ax13.set_ylim(0, 1.2)
style_axes(ax13)
ax13.legend(prop=FONT_PROPS)
save_plot(fig13, "14_dae_defense_effect.png")


# ============================================================
# STEP 21: HHO CONVERGENCE CURVE - FIGURE
# ============================================================

print("\n" + "=" * 70)
print("STEP 21: HHO CONVERGENCE CURVE")
print("=" * 70)

fig_hho, ax_hho = plt.subplots(figsize=(10, 8))

hho_iterations = list(range(1, len(HHO_CONVERGENCE_HISTORY) + 1))

ax_hho.plot(hho_iterations, HHO_CONVERGENCE_HISTORY, color=DARK_COLORS[0], linewidth=2.5, marker="o", markersize=6)

ax_hho.set_title("HHO Convergence Curve", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax_hho.set_xlabel("Iteration", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax_hho.set_ylabel("Best Validation Accuracy", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
style_axes(ax_hho)
save_plot(fig_hho, "15_hho_convergence_curve.png")


# ============================================================
# STEP 22: ABLATION STUDY - REPORT ONLY
# ============================================================

print("\n" + "=" * 70)
print("STEP 22: ABLATION STUDY")
print("=" * 70)

with torch.no_grad():
    denoised_clean_t = dae(X_test_tensor)
    blended_clean_t = BEST_BETA * denoised_clean_t + (1 - BEST_BETA) * X_test_tensor
    branch2_clean_prob_t = torch.sigmoid(robust_model(blended_clean_t))
    final_clean_prob_t = BEST_ALPHA * robust_clean_probs_t + (1 - BEST_ALPHA) * branch2_clean_prob_t
    final_clean_preds_t = (final_clean_prob_t > 0.5).float()

final_clean_preds_np = final_clean_preds_t.detach().cpu().numpy()
final_adv_preds_np_ablation = final_adv_preds.detach().cpu().numpy()
final_adv_probs_np_ablation = final_adv_prob.detach().cpu().numpy()

flipped_hybrid_mask = robust_clean_correct_mask & (final_adv_preds != y_test_tensor)
ASR_HYBRID = (flipped_hybrid_mask.sum().item() / num_robust_clean_correct) if num_robust_clean_correct > 0 else 0.0

ablation_results = pd.DataFrame({
    "Baseline (No Defense)": {
        "clean_accuracy": baseline_clean_metrics["accuracy"],
        "robust_accuracy": BASELINE_ROBUST_ACCURACY,
        "f1_adv": baseline_fgsm_metrics["f1"],
        "roc_auc_adv": baseline_fgsm_metrics["roc_auc"],
        "asr": ATTACK_SUCCESS_RATE,
    },
    "+ Adversarial Training": {
        "clean_accuracy": accuracy_score(y_test_orig_np, robust_clean_preds_t.detach().cpu().numpy()),
        "robust_accuracy": with_adv_training_metrics["robust_accuracy"],
        "f1_adv": with_adv_training_metrics["f1"],
        "roc_auc_adv": with_adv_training_metrics["roc_auc"],
        "asr": with_adv_training_metrics["asr"],
    },
    "+ DAE Only": {
        "clean_accuracy": accuracy_score(y_test_orig_np, (branch2_clean_prob_t > 0.5).float().detach().cpu().numpy()),
        "robust_accuracy": with_dae_metrics["accuracy"],
        "f1_adv": with_dae_metrics["f1"],
        "roc_auc_adv": roc_auc_score(y_test_orig_np, branch2_adv_probs_np),
        "asr": with_dae_metrics["asr"],
    },
    "Full Hybrid (Adv Training + DAE + HHO)": {
        "clean_accuracy": accuracy_score(y_test_orig_np, final_clean_preds_np),
        "robust_accuracy": accuracy_score(y_test_orig_np, final_adv_preds_np_ablation),
        "f1_adv": f1_score(y_test_orig_np, final_adv_preds_np_ablation, zero_division=0),
        "roc_auc_adv": roc_auc_score(y_test_orig_np, final_adv_probs_np_ablation),
        "asr": ASR_HYBRID,
    },
}).T

print("\n", ablation_results.round(5))


# ============================================================
# STEP 23: ROBUSTNESS UNDER DIFFERENT FGSM STRENGTHS - REPORT + FIGURE
# ============================================================

print("\n" + "=" * 70)
print("STEP 23: ROBUSTNESS UNDER DIFFERENT FGSM STRENGTHS")
print("=" * 70)

EPSILON_VALUES = [0.001, 0.005, 0.01, 0.02, 0.05]
epsilon_results = []

for eps_val in EPSILON_VALUES:

    X_eps_input = X_test_tensor.detach().clone().requires_grad_(True)

    eps_outputs = robust_model(X_eps_input)
    eps_loss = criterion(eps_outputs, y_test_tensor)

    robust_model.zero_grad()
    eps_loss.backward()

    eps_grad = X_eps_input.grad.data
    eps_perturbation = eps_val * eps_grad.sign()

    X_eps_adv = torch.clamp(X_eps_input.detach() + eps_perturbation, FEATURE_MIN, FEATURE_MAX)

    with torch.no_grad():

        eps_branch1_prob = torch.sigmoid(robust_model(X_eps_adv))

        eps_denoised = dae(X_eps_adv)
        eps_blended = BEST_BETA * eps_denoised + (1 - BEST_BETA) * X_eps_adv
        eps_branch2_prob = torch.sigmoid(robust_model(eps_blended))

        eps_final_prob = BEST_ALPHA * eps_branch1_prob + (1 - BEST_ALPHA) * eps_branch2_prob
        eps_final_preds = (eps_final_prob > 0.5).float()

    eps_final_preds_np = eps_final_preds.detach().cpu().numpy()

    eps_flipped_mask = robust_clean_correct_mask & (eps_final_preds != y_test_tensor)
    eps_asr = (eps_flipped_mask.sum().item() / num_robust_clean_correct) if num_robust_clean_correct > 0 else 0.0

    eps_accuracy = accuracy_score(y_test_orig_np, eps_final_preds_np)
    eps_f1 = f1_score(y_test_orig_np, eps_final_preds_np, zero_division=0)

    epsilon_results.append({
        "epsilon": eps_val,
        "robust_accuracy": eps_accuracy,
        "f1": eps_f1,
        "asr": eps_asr,
    })

    print(
        "Epsilon:", eps_val,
        "Robust Accuracy:", round(eps_accuracy, 5),
        "F1-score:", round(eps_f1, 5),
        "ASR:", round(eps_asr, 5)
    )

epsilon_df = pd.DataFrame(epsilon_results)
print("\n", epsilon_df.round(5))

fig14, ax14 = plt.subplots(figsize=(10, 8))

ax14.plot(epsilon_df["epsilon"], epsilon_df["robust_accuracy"], marker="o", color=DARK_COLORS[0], linewidth=2.5, markersize=8, label="Robust Accuracy")
ax14.plot(epsilon_df["epsilon"], epsilon_df["f1"], marker="s", color=DARK_COLORS[4], linewidth=2.5, markersize=8, label="F1-score")

ax14.set_title("Robust Accuracy vs FGSM Perturbation Strength", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax14.set_xlabel("Epsilon", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax14.set_ylabel("Score", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
style_axes(ax14)
ax14.legend(prop=FONT_PROPS)
save_plot(fig14, "16_robustness_vs_epsilon.png")

print("\nAll additional plots saved to folder:", os.path.abspath(PLOT_DIR))


# ============================================================
# STEP 24: MODEL COMPARISON (PROPOSED HYBRID VS BASELINE VS CNN VS LSTM)
# ============================================================

print("\n" + "=" * 70)
print("STEP 24: MODEL COMPARISON")
print("=" * 70)

with torch.no_grad():
    baseline_model_probs_c = torch.sigmoid(model(X_test_c_tensor)).cpu().numpy()

baseline_model_preds_c = (baseline_model_probs_c > 0.5).astype(float)
cnn_preds_c = (cnn_probs > 0.5).astype(float)
lstm_preds_c = (lstm_probs > 0.5).astype(float)

MODEL_COMPARISON = {
    "Baseline TabTransformer": {
        "accuracy": accuracy_score(y_test_np, baseline_model_preds_c),
        "precision": precision_score(y_test_np, baseline_model_preds_c, zero_division=0),
        "recall": recall_score(y_test_np, baseline_model_preds_c, zero_division=0),
        "f1": f1_score(y_test_np, baseline_model_preds_c, zero_division=0),
        "roc_auc": roc_auc_score(y_test_np, baseline_model_probs_c),
    },
    "CNN": {
        "accuracy": accuracy_score(y_test_np, cnn_preds_c),
        "precision": precision_score(y_test_np, cnn_preds_c, zero_division=0),
        "recall": recall_score(y_test_np, cnn_preds_c, zero_division=0),
        "f1": f1_score(y_test_np, cnn_preds_c, zero_division=0),
        "roc_auc": cnn_auc,
    },
    "LSTM": {
        "accuracy": accuracy_score(y_test_np, lstm_preds_c),
        "precision": precision_score(y_test_np, lstm_preds_c, zero_division=0),
        "recall": recall_score(y_test_np, lstm_preds_c, zero_division=0),
        "f1": f1_score(y_test_np, lstm_preds_c, zero_division=0),
        "roc_auc": lstm_auc,
    },
    "Proposed Hybrid": {
        "accuracy": MODEL_RESULTS["Final Hybrid"]["accuracy"],
        "precision": MODEL_RESULTS["Final Hybrid"]["precision"],
        "recall": MODEL_RESULTS["Final Hybrid"]["recall"],
        "f1": MODEL_RESULTS["Final Hybrid"]["f1"],
        "roc_auc": MODEL_RESULTS["Final Hybrid"]["roc_auc"],
    },
}

model_comparison_df = pd.DataFrame(MODEL_COMPARISON).T

print("\nModel Comparison (Proposed Hybrid vs Baseline TabTransformer vs CNN vs LSTM):")
print(model_comparison_df.round(5))

fig15, ax15 = plt.subplots(figsize=(16, 10))

comparison_metric_labels = ["Accuracy", "Precision", "Recall", "F1-score", "ROC-AUC"]
comparison_metric_keys = ["accuracy", "precision", "recall", "f1", "roc_auc"]

model_names_comparison = list(MODEL_COMPARISON.keys())
n_models_comparison = len(model_names_comparison)
x_pos_comp = np.arange(len(comparison_metric_labels))
comp_bar_width = 0.8 / n_models_comparison

comparison_colors = [DARK_COLORS[1], DARK_COLORS[3], DARK_COLORS[5], DARK_COLORS[0]]

for idx, model_name in enumerate(model_names_comparison):
    values = [MODEL_COMPARISON[model_name][k] for k in comparison_metric_keys]
    offset = (idx - (n_models_comparison - 1) / 2) * comp_bar_width
    bars = ax15.bar(
        x_pos_comp + offset, values, comp_bar_width,
        color=comparison_colors[idx % len(comparison_colors)],
        label=model_name
    )
    annotate_bars(ax15, bars)

ax15.set_title("Model Comparison", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax15.set_xlabel("Metric", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax15.set_ylabel("Score", fontweight="bold", fontfamily="Times New Roman", fontsize=FONT_SIZE)
ax15.set_xticks(x_pos_comp)
ax15.set_xticklabels(comparison_metric_labels)
ax15.set_ylim(0, 1.15)
style_axes(ax15)
ax15.legend(prop=FONT_PROPS)
save_plot(fig15, "17_model_comparison.png")


# ============================================================
# STEP 25: EXPORT REPORT TO EXCEL
# ============================================================

print("\n" + "=" * 70)
print("STEP 25: EXPORT REPORT TO EXCEL")
print("=" * 70)

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

REPORT_PATH = "IDS_Pipeline_Report.xlsx"

feature_selection_df = pd.DataFrame([{
    "Original Features": N_ORIGINAL_FEATURES,
    "Selected Features": N_SELECTED_FEATURES,
    "Removed Features": N_REMOVED_FEATURES,
    "Feature Reduction (%)": round(PCT_FEATURE_REDUCTION, 2),
}])

fgsm_impact_df = pd.DataFrame({
    "Clean": baseline_clean_metrics,
    "FGSM": baseline_fgsm_metrics,
}).T
fgsm_impact_df["accuracy_degradation"] = [0.0, round(BASELINE_ACCURACY_DEGRADATION, 5)]
fgsm_impact_df["asr"] = [None, round(ATTACK_SUCCESS_RATE, 5)]
fgsm_impact_df["robust_accuracy"] = [None, round(BASELINE_ROBUST_ACCURACY, 5)]

dae_effect_df = pd.DataFrame({
    "Without DAE": without_dae_metrics,
    "With DAE": with_dae_metrics,
}).T

hho_history_df = pd.DataFrame({
    "Iteration": list(range(1, len(HHO_CONVERGENCE_HISTORY) + 1)),
    "Best Validation Accuracy": HHO_CONVERGENCE_HISTORY,
})

final_metrics_df = pd.DataFrame({
    name: {
        "accuracy": res["accuracy"],
        "precision": res["precision"],
        "recall": res["recall"],
        "f1": res["f1"],
        "roc_auc": res["roc_auc"],
        "fpr": res["fpr"],
        "fnr": res["fnr"],
    }
    for name, res in MODEL_RESULTS.items()
}).T

baseline_comparison_df = pd.DataFrame({
    "Proposed Hybrid (ROC-AUC)": [MODEL_RESULTS["Final Hybrid"]["roc_auc"]],
    "CNN (ROC-AUC)": [cnn_auc],
    "LSTM (ROC-AUC)": [lstm_auc],
})

report_sheets = {
    "Feature Selection": feature_selection_df,
    "FGSM Attack Impact": fgsm_impact_df.round(5),
    "Adversarial Training": adv_training_comparison.round(5),
    "DAE Defense Effect": dae_effect_df.round(5),
    "Ablation Study": ablation_results.round(5),
    "Epsilon Robustness": epsilon_df.round(5),
    "HHO Convergence": hho_history_df.round(5),
    "Final Model Metrics": final_metrics_df.round(5),
    "Baseline Model Comparison": baseline_comparison_df.round(5),
    "Model Comparison": model_comparison_df.round(5),
}

with pd.ExcelWriter(REPORT_PATH, engine="openpyxl") as writer:
    for sheet_name, df in report_sheets.items():
        df.to_excel(writer, sheet_name=sheet_name[:31], index=True)

workbook = load_workbook(REPORT_PATH)

for sheet in workbook.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = Font(name="Times New Roman", bold=(cell.row == 1))
            cell.alignment = Alignment(horizontal="center")
    for column_cells in sheet.columns:
        max_length = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells)
        col_letter = column_cells[0].column_letter
        sheet.column_dimensions[col_letter].width = max(12, max_length + 2)

workbook.save(REPORT_PATH)

print("\nExcel report saved to:", os.path.abspath(REPORT_PATH))

plt.show()